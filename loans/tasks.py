from celery import shared_task
import openpyxl
from decimal import Decimal
from datetime import datetime
from django.db import transaction
from .models import Customer, Loan
import logging

logger = logging.getLogger(__name__)


@shared_task
def ingest_customer_data(file_path):
    """
    Background task to ingest customer data from Excel file
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        created_count = 0
        updated_count = 0
        
        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Check if customer_id exists
                    customer_id, first_name, last_name, age, phone_number, monthly_salary, approved_limit = row
                    
                    customer, created = Customer.objects.update_or_create(
                        customer_id=customer_id,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'age': age,
                            'phone_number': str(phone_number),
                            'monthly_salary': Decimal(str(monthly_salary)),
                            'approved_limit': Decimal(str(approved_limit)),
                            'current_debt': Decimal('0')  # Set to 0 initially, will calculate later
                        }
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
        
        logger.info(f"Customer data ingestion completed. Created: {created_count}, Updated: {updated_count}")
        return f"Customer data ingestion completed. Created: {created_count}, Updated: {updated_count}"
        
    except Exception as e:
        logger.error(f"Error ingesting customer data: {str(e)}")
        return f"Error ingesting customer data: {str(e)}"


@shared_task
def ingest_loan_data(file_path):
    """
    Background task to ingest loan data from Excel file
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        created_count = 0
        
        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # Check if customer_id and loan_id exist
                    customer_id, loan_id, loan_amount, tenure, interest_rate, monthly_repayment, emis_paid_on_time, start_date, end_date = row
                    
                    try:
                        customer = Customer.objects.get(customer_id=customer_id)
                        
                        # Convert date strings to date objects if they're strings
                        if isinstance(start_date, str):
                            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                        if isinstance(end_date, str):
                            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                        
                        loan, created = Loan.objects.update_or_create(
                            loan_id=loan_id,
                            defaults={
                                'customer': customer,
                                'loan_amount': Decimal(str(loan_amount)),
                                'tenure': int(float(tenure)),
                                'interest_rate': Decimal(str(interest_rate)),
                                'monthly_repayment': Decimal(str(monthly_repayment)),
                                'emis_paid_on_time': int(float(emis_paid_on_time) if emis_paid_on_time else 0),
                                'start_date': start_date,
                                'end_date': end_date,
                            }
                        )
                        
                        if created:
                            created_count += 1
                            
                    except Customer.DoesNotExist:
                        logger.warning(f"Customer {customer_id} not found for loan {loan_id}")
                        continue
        
        logger.info(f"Loan data ingestion completed. Created: {created_count}")
        return f"Loan data ingestion completed. Created: {created_count}"
        
    except Exception as e:
        logger.error(f"Error ingesting loan data: {str(e)}")
        return f"Error ingesting loan data: {str(e)}"


@shared_task
def update_customer_current_debt():
    """
    Background task to update current debt for all customers based on active loans
    """
    try:
        customers = Customer.objects.all()
        updated_count = 0
        
        with transaction.atomic():
            for customer in customers:
                # Calculate current debt from active loans
                active_loans = Loan.objects.filter(
                    customer=customer,
                    end_date__gte=datetime.now().date()
                )
                
                current_debt = sum(loan.loan_amount for loan in active_loans)
                customer.current_debt = current_debt
                customer.save()
                updated_count += 1
        
        logger.info(f"Updated current debt for {updated_count} customers")
        return f"Updated current debt for {updated_count} customers"
        
    except Exception as e:
        logger.error(f"Error updating current debt: {str(e)}")
        return f"Error updating current debt: {str(e)}"
